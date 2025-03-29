import discord
from discord.ext import commands
import os
from openpyxl import Workbook, load_workbook
from dotenv import load_dotenv
import requests
from discord import Option, Attachment
import math
from io import BytesIO
import json

# Wczytanie tokenu z pliku .env
load_dotenv("Token.env")
TOKEN = os.getenv('DISCORD_TOKEN')
YOUR_USER_ID = 311552371970932736
YOUR_CHANNEL_ID = 1347931852157227089  # Zastąp prawdziwym ID kanału
IMGUR_CLIENT_ID = os.getenv('IMGUR_CLIENT_ID')
IMGUR_UPLOAD_URL = "https://api.imgur.com/3/upload"
SUBMISSIONS_FILE = "submissions.json"

# Globalna wartość wersji
global_version = "47.3"  # Domyślna wartość

# Słowniki (do uzupełnienia)
towers = {
    "dart#000": 200,
    "dart#100": 140,
    "dart#200": 200,
    "dart#300": 320,
    "dart#400": 1800,
    "dart#500": 15000,
    "dart#010": 100,
    "dart#020": 190,
    "dart#030": 450,
    "dart#040": 7200,
    "dart#050": 45000,
    "dart#001": 90,
    "dart#002": 200,
    "dart#003": 575,
    "dart#004": 2050,
    "dart#005": 21500,
    "boomer#000": 315,
    "boomer#100": 200,
    "boomer#200": 280,
    "boomer#300": 600,
    "boomer#400": 3000,
    "boomer#500": 32500,
    "boomer#010": 175,
    "boomer#020": 250,
    "boomer#030": 1250,
    "boomer#040": 4200,
    "boomer#050": 35000,
    "boomer#001": 100,
    "boomer#002": 300,
    "boomer#003": 1300,
    "boomer#004": 2400,
    "boomer#005": 50000,
    "bomb#000": 375,
    "bomb#100": 250,
    "bomb#200": 650,
    "bomb#300": 1100,
    "bomb#400": 2800,
    "bomb#500": 55000,
    "bomb#010": 250,
    "bomb#020": 400,
    "bomb#030": 1000,
    "bomb#040": 3450,
    "bomb#050": 28000,
    "bomb#001": 200,
    "bomb#002": 300,
    "bomb#003": 700,
    "bomb#004": 2500,
    "bomb#005": 23000,
    "tack#000": 260,
    "tack#100": 150,
    "tack#200": 300,
    "tack#300": 600,
    "tack#400": 3500,
    "tack#500": 45500,
    "tack#010": 100,
    "tack#020": 225,
    "tack#030": 550,
    "tack#040": 2700,
    "tack#050": 15000,
    "tack#001": 110,
    "tack#002": 110,
    "tack#003": 450,
    "tack#004": 3200,
    "tack#005": 20000,
    "ice#000": 500,
    "ice#100": 150,
    "ice#200": 350,
    "ice#300": 1500,
    "ice#400": 2200,
    "ice#500": 28000,
    "ice#010": 225,
    "ice#020": 450,
    "ice#030": 2800,
    "ice#040": 3800,
    "ice#050": 19200,
    "ice#001": 175,
    "ice#002": 225,
    "ice#003": 1750,
    "ice#004": 2750,
    "ice#005": 30000,
    "glue#000": 225,
    "glue#100": 200,
    "glue#200": 300,
    "glue#300": 2000,
    "glue#400": 5000,
    "glue#500": 22500,
    "glue#010": 100,
    "glue#020": 970, 
    "glue#030": 1950,
    "glue#040": 4000,
    "glue#050": 16000,
    "glue#001": 280,
    "glue#002": 400,
    "glue#003": 3600,
    "glue#004": 4000,
    "glue#005": 24000,
    "sniper#000": 350,
    "sniper#100": 350,
    "sniper#200": 1300,
    "sniper#300": 3000,
    "sniper#400": 5650,
    "sniper#500": 32000,
    "sniper#010": 250,
    "sniper#020": 450,
    "sniper#030": 2100,
    "sniper#040": 7600,
    "sniper#050": 14500,
    "sniper#001": 450,
    "sniper#002": 450,
    "sniper#003": 2900,
    "sniper#004": 4100,
    "sniper#005": 14700,
    "sub#000": 325,
    "sub#100": 130,
    "sub#200": 500,
    "sub#300": 700,
    "sub#400": 2300,
    "sub#500": 28000,
    "sub#010": 450,
    "sub#020": 300,
    "sub#030": 1350,
    "sub#040": 13000,
    "sub#050": 29000,
    "sub#001": 450,
    "sub#002": 1000,
    "sub#003": 1100,
    "sub#004": 2500,
    "sub#005": 25000,
    "bucc#000": 400,
    "bucc#100": 275,
    "bucc#200": 425,
    "bucc#300": 3050,
    "bucc#400": 8000,
    "bucc#500": 24500,
    "bucc#010": 550,
    "bucc#020": 500,
    "bucc#030": 900,
    "bucc#040": 3900,
    "bucc#050": 27000,
    "bucc#001": 200,
    "bucc#002": 350,
    "bucc#003": 2400,
    "bucc#004": 5500,
    "bucc#005": 23000,
    "ace#000": 800,
    "ace#100": 650,
    "ace#200": 650,
    "ace#300": 1000,
    "ace#400": 3000,
    "ace#500": 42500,
    "ace#010": 200,
    "ace#020": 350,
    "ace#030": 900,
    "ace#040": 18000,
    "ace#050": 30000,
    "ace#001": 500,
    "ace#002": 550,
    "ace#003": 2550,
    "ace#004": 23400,
    "ace#005": 85000,
    "heli#000": 1600,
    "heli#100": 800,
    "heli#200": 500,
    "heli#300": 1750,
    "heli#400": 19600,
    "heli#500": 45000,
    "heli#010": 300,
    "heli#020": 600,
    "heli#030": 3500,
    "heli#040": 9500,
    "heli#050": 30000,
    "heli#001": 250,
    "heli#002": 350,
    "heli#003": 3000,
    "heli#004": 8500,
    "heli#005": 35000,
    "mortar#000": 750,
    "mortar#100": 500,
    "mortar#200": 500,
    "mortar#300": 825,
    "mortar#400": 7200,
    "mortar#500": 36000,
    "mortar#010": 300,
    "mortar#020": 500,
    "mortar#030": 900,
    "mortar#040": 6500,
    "mortar#050": 38000,
    "mortar#001": 200,
    "mortar#002": 500,
    "mortar#003": 900,
    "mortar#004": 9500,
    "mortar#005": 40000,
    "dartling#000": 850,
    "dartling#100": 300,
    "dartling#200": 900,
    "dartling#300": 3000,
    "dartling#400": 11750,
    "dartling#500": 75000,
    "dartling#010": 250,
    "dartling#020": 950,
    "dartling#030": 4500,
    "dartling#040": 5850,
    "dartling#050": 65000,
    "dartling#001": 150,
    "dartling#002": 1200,
    "dartling#003": 3400,
    "dartling#004": 12000,
    "dartling#005": 58000,
    "wiz#000": 325,
    "wiz#100": 150,
    "wiz#200": 450,
    "wiz#300": 1400,
    "wiz#400": 10000,
    "wiz#500": 32000,
    "wiz#010": 300,
    "wiz#020": 800,
    "wiz#030": 3300,
    "wiz#040": 6000,
    "wiz#050": 50000,
    "wiz#001": 300,
    "wiz#002": 300,
    "wiz#003": 1500,
    "wiz#004": 2800,
    "wiz#005": 26500,
    "super#000": 2500,
    "super#100": 2000,
    "super#200": 2500,
    "super#300": 20000,
    "super#400": 100000,
    "super#500": 500000,
    "super#010": 1500,
    "super#020": 1900,
    "super#030": 7500,
    "super#040": 25000,
    "super#050": 70000,
    "super#001": 3000,
    "super#002": 1200,
    "super#003": 5600,
    "super#004": 55555,
    "super#005": 165650,
    "ninja#000": 400,
    "ninja#100": 350,
    "ninja#200": 350,
    "ninja#300": 900,
    "ninja#400": 2750,
    "ninja#500": 35000,
    "ninja#010": 250,
    "ninja#020": 400,
    "ninja#030": 1200,
    "ninja#040": 5200,
    "ninja#050": 22000,
    "ninja#001": 300,
    "ninja#002": 450,
    "ninja#003": 2250,
    "ninja#004": 5000,
    "ninja#005": 40000,
    "alch#000": 550,
    "alch#100": 250,
    "alch#200": 350,
    "alch#300": 1400,
    "alch#400": 2850,
    "alch#500": 48000,
    "alch#010": 250,
    "alch#020": 475,
    "alch#030": 2800,
    "alch#040": 4200,
    "alch#050": 45000,
    "alch#001": 650,
    "alch#002": 450,
    "alch#003": 1000,
    "alch#004": 2750,
    "alch#005": 40000,
    "druid#000": 400,
    "druid#100": 350,
    "druid#200": 850,
    "druid#300": 1700,
    "druid#400": 4500,
    "druid#500": 60000,
    "druid#010": 250,
    "druid#020": 350,
    "druid#030": 1050,
    "druid#040": 4900,
    "druid#050": 35000,
    "druid#001": 100,
    "druid#002": 300,
    "druid#003": 600,
    "druid#004": 2350,
    "druid#005": 45000,
    "farm#000": 1250,
    "farm#100": 500,
    "farm#200": 600,
    "farm#300": 3000,
    "farm#400": 19000,
    "farm#500": 115000,
    "farm#010": 300,
    "farm#020": 800,
    "farm#030": 3650,
    "farm#040": 7200,
    "farm#050": 100000,
    "farm#001": 250,
    "farm#002": 400,
    "farm#003": 2700,
    "farm#004": 15000,
    "farm#005": 70000,
    "spac#000": 1000,
    "spac#100": 800,
    "spac#200": 600,
    "spac#300": 2300,
    "spac#400": 9500,
    "spac#500": 125000,
    "spac#010": 600,
    "spac#020": 800,
    "spac#030": 2500,
    "spac#040": 6000,
    "spac#050": 42000,
    "spac#001": 150,
    "spac#002": 400,
    "spac#003": 1300,
    "spac#004": 3600,
    "spac#005": 30000,
    "vill#000": 1200,
    "vill#100": 400,
    "vill#200": 1500,
    "vill#300": 800,
    "vill#400": 2500,
    "vill#500": 25000,
    "vill#010": 250,
    "vill#020": 2000,
    "vill#030": 7500,
    "vill#040": 20000,
    "vill#050": 40000,
    "vill#001": 500,
    "vill#002": 500,
    "vill#003": 10000,
    "vill#004": 3000,
    "vill#005": 5000,
    "engi#000": 350,
    "engi#100": 500,
    "engi#200": 400,
    "engi#300": 575,
    "engi#400": 2500,
    "engi#500": 32000,
    "engi#010": 250,
    "engi#020": 350,
    "engi#030": 900,
    "engi#040": 13500,
    "engi#050": 72000,
    "engi#001": 450,
    "engi#002": 220,
    "engi#003": 450,
    "engi#004": 3600,
    "engi#005": 45000,
    "beast#000": 250,
    "beast#100": 160,
    "beast#200": 810,
    "beast#300": 2010,
    "beast#400": 12500,
    "beast#500": 45000,
    "beast#010": 175,
    "beast#020": 830,
    "beast#030": 2065,
    "beast#040": 9500,
    "beast#050": 60000,
    "beast#001": 190,
    "beast#002": 860,
    "beast#003": 2120,
    "beast#004": 9000,
    "beast#005": 30000,
    "merm#000": 475,
    "merm#100": 250,
    "merm#200": 300,
    "merm#300": 2100,
    "merm#400": 3600,
    "merm#500": 23000,
    "merm#010": 300,
    "merm#020": 400,
    "merm#030": 2300,
    "merm#040": 8000,
    "merm#050": 52000,
    "merm#001": 200,
    "merm#002": 380,
    "merm#003": 2000,
    "merm#004": 7600,
    "merm#005": 25000
}

heroes = {
    "Adora": {
        "cost": 1000,
        "levelModifier": 1.71
    },
    "Benjamin": {
        "cost": 1200,
        "levelModifier": 1.5
    },
    "Brickell": {
        "cost": 900,
        "levelModifier": 1.425
    },
    "Churchill": {
        "cost": 2000,
        "levelModifier": 1.71
    },
    "Corvus": {
        "cost": 1025,
        "levelModifier": 1.425
    },
    "Etienne": {
        "cost": 850,
        "levelModifier": 1
    },
    "Ezili": {
        "cost": 600,
        "levelModifier": 1.425
    },
    "Geraldo": {
        "cost": 750,
        "levelModifier": 1
    },
    "Gwen": {
        "cost": 725,
        "levelModifier": 1
    },
    "Jones": {
        "cost": 700,
        "levelModifier": 1
    },
    "Obyn": {
        "cost": 650,
        "levelModifier": 1
    },
    "Pat": {
        "cost": 800,
        "levelModifier": 1.425
    },
    "Psi": {
        "cost": 1000,
        "levelModifier": 1.5
    },
    "Quincy": {
        "cost": 540,
        "levelModifier": 1
    },
    "Sauda": {
        "cost": 600,
        "levelModifier": 1.425
    },
    "Rosalia": {
        "cost": 875,
        "levelModifier": 1.425
    }
}

hero_level_costs = [
    0,      # Poziom 1 (koszt bazowy)
    180,    # Poziom 2
    460,    # Poziom 3
    1000,   # Poziom 4
    1860,   # Poziom 5
    3280,   # Poziom 6
    5180,   # Poziom 7
    8320,   # Poziom 8
    9380,   # Poziom 9
    13620,  # Poziom 10
    16380,  # Poziom 11
    14400,  # Poziom 12
    16650,  # Poziom 13
    14940,  # Poziom 14
    16380,  # Poziom 15
    17820,  # Poziom 16
    19260,  # Poziom 17
    20700,  # Poziom 18
    16470,  # Poziom 19
    17280   # Poziom 20
]


maps = {
    "Logs": "Logs",
    "Resort": "Resort",
    "Cubism": "Cubism",
    "Candy Falls": "Candy Falls",
    "Carved": "Carved",
    "In the Loop": "In the Loop",
    "Tree Stump": "Tree Stump",
    "Hedge": "Hedge",
    "Skates": "Skates",
    "Four Circles": "Four Circles",
    "Winter Park": "Winter Park",
    "Frozen Over": "Frozen Over",
    "Park Path": "Park Path",
    "Lotus Island": "Lotus Island",
    "Alpine Run": "Alpine Run",
    "Town Center": "Town Center",
    "Monkey Meadow": "Monkey Meadow",
    "End Of The Road": "End Of The Road",
    "The Cabin": "The Cabin",
    "Scrapyard": "The Cabin",
    "One Two Tree": "One Two Tree",
    "Middle Of The Road": "Middle Of The Road",
    "Tinkerton": "Tinkerton",


    "Moon Landing": "Moon Landing",
    "Cracked": "Cracked",
    "Kartsndarts": "Kartsndarts",
    "Firing Range": "Firing Range",
    "Adora'S Temple": "Adora'S Temple",
    "Downstream": "Downstream",
    "Streambed": "Streambed",
    "Spring Spring": "Spring Spring",
    "Spice Islands": "Spice Islands",
    "Bazaar": "Bazaar",
    "Haunted": "Haunted",
    "Chutes": "Chutes",
    "Rake": "Rake",
    "Encrypted": "Encrypted",
    "Balance": "Balance",
    "Bloonarius Prime": "Bloonarius Prime",
    "Quiet Street": "Quiet Street",
    "Quarry": "Quarry",
    "Covered Garden": "Covered Garden",
    "Polyphemus": "Polyphemus",
    "Water Park": "Water Park",
    "Sulfur Springs": "Sulfur Springs",
    "Protect The Yacht": "Protect The Yacht",
    "Luminous Cove": "Luminous Cove",


    "Cornfield": "Cornfield",
    "Spillway": "Spillway",
    "High Finance": "High Finance",
    "Another Brick": "Another Brick",
    "Off The Coast": "Off The Coast",
    "Peninsula": "Peninsula",
    "Pat'S Pond": "Pat'S Pond",
    "Geared": "Geared",
    "Underground": "Underground",
    "Cargo": "Cargo",
    "Mesa": "Mesa",
    "X Factor": "X Factor",
    "Sunken Columns": "Sunken Columns",
    "Midnight Mansion": "Midnight Mansion",
    "Erosion": "Erosion",
    "Dark Path": "Dark Path",
    "Castle Revenge": "Castle Revenge",
    "Ancient Portal": "Ancient Portal",
    "Last Resort": "Last Resort",
    "Enchanted Glade": "Enchanted Glade",


    "Infernal": "Infernal",
    "Workshop": "Workshop",
    "Quad": "Quad",
    "Muddy Puddles": "Muddy Puddles",
    "#Ouch": "#Ouch",
    "Bloody Puddles": "Bloody Puddles",
    "Flooded Valley": "Flooded Valley",
    "Dark Castle": "Dark Castle",
    "Ravine": "Ravine",
    "Blons": "Blons",
    "Sanctuary": "Sanctuary",
    "Dark Dungeons": "Dark Dungeons",
    "Glacial Trail": "Glacial Trail"
}

maps_with_aliases = {
    'Logs': {
        'aliases': ['l', 'log'],
        'value': 'Logs'
    },
    'Resort': {
        'aliases': ['rs', 'res', 'reso'],
        'value': 'Resort'
    },
    'Cubism': {
        'aliases': ['c', 'cub', 'cube', 'cubi', 'cubism'],
        'value': 'Cubism'
    },
    'Candy Falls': {
        'aliases': ['cy', 'candy', 'cfalls', 'falls'],
        'value': 'Candy Falls'
    },
    'Carved': {
        'aliases': ['cv', 'cvd', 'carve'],
        'value': 'Carved'
    },
    'In The Loop': {
        'aliases': ['il', 'itl', 'loop', 'in', 'in_the'],
        'value': 'In The Loop'
    },
    'Tree Stump': {
        'aliases': ['ts', 'tree', 'stump'],
        'value': 'Tree Stump'
    },
    'Hedge': {
        'aliases': ['h', 'hdg', 'hej'],
        'value': 'Hedge'
    },
    'Skates': {
        'aliases': ['s', 'skate', 'sk'],
        'value': 'Skates'
    },
    'Four Circles': {
        'aliases': ['fc', '4c', 'circle', 'four'],
        'value': 'Four Circles'
    },
    'Winter Park': {
        'aliases': ['wp', 'winter'],
        'value': 'Winter Park'
    },
    'Frozen Over': {
        'aliases': ['fo', 'frozen', 'frover'],
        'value': 'Frozen Over'
    },
    'Park Path': {
        'aliases': ['pk', 'park', 'ppath'],
        'value': 'Park Path'
    },
    'Lotus Island': {
        'aliases': ['li', 'lotus', 'lot', 'island'],
        'value': 'Lotus Island'
    },
    'Alpine Run': {
        'aliases': ['ar', 'alpine', 'alp', 'run'],
        'value': 'Alpine Run'
    },
    'Town Center': {
        'aliases': ['tc', 'town', 'center'],
        'value': 'Town Center'
    },
    'Monkey Meadow': {
        'aliases': ['mm', 'm_meadow', 'meadow', 'monkey'],
        'value': 'Monkey Meadow'
    },
    'End Of The Road': {
        'aliases': ['er', 'end', 'eotr'],
        'value': 'End Of The Road'
    },
    'The Cabin': {
        'aliases': ['cb', 'cbt', 'cab', 'cabin', 'le_cabin'],
        'value': 'The Cabin'
    },
    'Scrapyard': {
        'aliases': ['sy', 'scrap', 'scr', 'scraps', 'yard', 'syard', 'scyard'],
        'value': 'Scrapyard'
    },
    'One Two Tree': {
        'aliases': ['ott', '123'],
        'value': 'One Two Tree'
    },
    'Middle Of The Road': {
        'aliases': ['mr', 'motr', 'mother'],
        'value': 'Middle Of The Road'
    },
    'Tinkerton': {
        'aliases': ['tk', 'tink', 'tinker'],
        'value': 'Tinkerton'
    },



    'Moon Landing': {
        'aliases': ['ml', 'moon', 'landing'],
        'value': 'Moon Landing'
    },
    'Cracked': {
        'aliases': ['cr', 'ck', 'crkd', 'crack'],
        'value': 'Cracked'
    },
    'Kartsndarts': {
        'aliases': ['kd', 'kar', 'kart', 'karts', 'karts_n_darts ', 'kartsn', 'knd'],
        'value': 'Kartsndarts'
    },
    'Firing Range': {
        'aliases': ['fr', 'fir', 'firing'],
        'value': 'Firing Range'
    },
    "Adora's Temple": {
        'aliases': ['at', 'atemple', "adora's"],
        'value': "Adora's Temple"
    },
    'Downstream': {
        'aliases': ['ds', 'down'],
        'value': 'Downstream'
    },
    'Streambed': {
        'aliases': ['sb', 'stream'],
        'value': 'Streambed'
    },
    'Spring Spring': {
        'aliases': ['ss', 'spring'],
        'value': 'Spring Spring'
    },
    'Spice Islands': {
        'aliases': ['sp', 'si', 'spice'],
        'value': 'Spice Islands'
    },
    'Bazaar': {
        'aliases': ['bz', 'baz'],
        'value': 'Bazaar'
    },
    'Haunted': {
        'aliases': ['ha', 'haunt'],
        'value': 'Haunted'
    },
    'Chutes': {
        'aliases': ['ch', 'chut'],
        'value': 'Chutes'
    },
    'Rake': {
        'aliases': ['r', 'rk', 'rke'],
        'value': 'Rake'
    },
    'Encrypted': {
        'aliases': ['en', 'crypt', 'encrypt'],
        'value': 'Encrypted'
    },
    'Balance': {
        'aliases': ['ba', 'balanced', 'bal'],
        'value': 'Balance'
    },
    'Bloonarius Prime': {
        'aliases': ['blp', 'bloonarius', 'bprime'],
        'value': 'Bloonarius Prime'
    },
    'Quiet Street': {
        'aliases': ['qs', 'quiet'],
        'value': 'Quiet Street'
    },
    'Quarry': {
        'aliases': ['qy'],
        'value': 'Quarry'
    },
    'Covered Garden': {
        'aliases': ['ga', 'covered', 'garden', 'cgar'],
        'value': 'Covered Garden'
    },
    'Polyphemus': {
        'aliases': ['py', 'poly'],
        'value': 'Polyphemus'
    },
    'Water Park': {
        'aliases': ['wa', 'wpark'],
        'value': 'Water Park'
    },
    'Sulfur Springs': {
        'aliases': ['su', 'sulfur', 'sulphur', 'sulphur_springs'],
        'value': 'Sulfur Springs'
    },
    'Protect The Yacht': {
        'aliases': ['yt', 'yacht', 'mr_beast'],
        'value': 'Protect The Yacht'
    },
    'Luminous Cove': {
        'aliases': ['lc', 'luminous', 'cove', 'lumi', 'lum', 'lcove'],
        'value': 'Luminous Cove'
    },




    'Cornfield': {
        'aliases': ['cf', 'corn', 'field'],
        'value': 'Cornfield'
    },
    'Spillway': {
        'aliases': ['sw', 'spill', 'way'],
        'value': 'Spillway'
    },
    'High Finance': {
        'aliases': ['hf', 'hi_fi', 'finance', 'fin', 'high'],
        'value': 'High Finance'
    },
    'Another Brick': {
        'aliases': ['ab', 'abrick', 'another'],
        'value': 'Another Brick'
    },
    'Off The Coast': {
        'aliases': ['of', 'off', 'otc', 'coast'],
        'value': 'Off The Coast'
    },
    'Peninsula': {
        'aliases': ['pn', 'peni', 'pen', 'penin'],
        'value': 'Peninsula'
    },
    "Pat's Pond": {
        'aliases': ['pp', 'pond', 'ppond', 'pats_pond', 'pats'],
        'value': "Pat's Pond"
    },
    'Geared': {
        'aliases': ['gd', 'gear', 'grd', 'geard'],
        'value': 'Geared'
    },
    'Underground': {
        'aliases': ['ug', 'under', 'ground'],
        'value': 'Underground'
    },
    'Cargo': {
        'aliases': ['cg', 'cgo', 'car', 'go'],
        'value': 'Cargo'
    },
    'Mesa': {
        'aliases': ['ms'],
        'value': 'Mesa'
    },
    'X Factor': {
        'aliases': ['xf', 'xfac', 'factor'],
        'value': 'X Factor'
    },
    'Sunken Columns': {
        'aliases': ['sc', 'sunk', 'sunken', 's_col'],
        'value': 'Sunken Columns'
    },
    'Midnight Mansion': {
        'aliases': ['mn', 'midnight', 'mansion'],
        'value': 'Midnight Mansion'
    },
    'Erosion': {
        'aliases': ['e', 'ero'],
        'value': 'Erosion'
    },
    'Dark Path': {
        'aliases': ['dp'],
        'value': 'Dark Path'
    },
    'Castle Revenge': {
        'aliases': ['ca', 'revenge'],
        'value': 'Castle Revenge'
    },
    'Ancient Portal': {
        'aliases': ['ap', 'ap_physics'],
        'value': 'Ancient Portal'
    },
    'Last Resort': {
        'aliases': ['lr', 'last', 'last_remote'],
        'value': 'Last Resort'
    },
    'Enchanted Glade': {
        'aliases': ['eg', 'glade', 'eglade'],
        'value': 'Enchanted Glade'
    },



    'Infernal': {
        'aliases': ['if', 'i', 'inf'],
        'value': 'Infernal'
    },
    'Workshop': {
        'aliases': ['w', 'ws', 'wk', 'work', 'shop'],
        'value': 'Workshop'
    },
    'Quad': {
        'aliases': ['qd', 'kwad'],
        'value': 'Quad'
    },
    'Muddy Puddles': {
        'aliases': ['mp', 'muddles', 'mpuddles', 'muddy', 'cuddles'],
        'value': 'Muddy Puddles'
    },
    '#Ouch': {
        'aliases': ['ou', '#o', 'o', 'couch', 'ouch'],
        'value': '#Ouch'
    },
    'Bloody Puddles': {
        'aliases': ['bp', 'bloodles', 'bloody'],
        'value': 'Bloody Puddles'
    },
    'Flooded Valley': {
        'aliases': ['fv', 'flooded', 'valley'],
        'value': 'Flooded Valley'
    },
    'Dark Castle': {
        'aliases': ['dc', 'dca', 'dank', 'dank_castle', 'castle'],
        'value': 'Dark Castle'
    },
    'Ravine': {
        'aliases': ['rv', 'rav'],
        'value': 'Ravine'
    },
    'Blons': {
        'aliases': ['bl', 'corner', 'slons'],
        'value': 'Blons'
    },
    'Sanctuary': {
        'aliases': ['sa', 'sanc'],
        'value': 'Sanctuary'
    },
    'Dark Dungeons': {
        'aliases': ['dd', 'dark', 'dungeon', 'dungeons', 'dark_dungeon', 'd&d', 'dnd'],
        'value': 'Dark Dungeons'
    },
    'Glacial Trail': {
        'aliases': ['gt', 'glac', 'glacier'],
        'value': 'Glacial Trail'
    }
}



# Słownik aliasów
aliases = {alias: value for aliases_group in 
           [({"dart", "dart_monkey", "drt", "dm"}, "dart"), 
            ({"boomer", "boomerang_monkey", "boomerang", "bm", "rang", "bomerang", "boo", "bomer", "rangs", "bomerrang"}, "boomer"), 
            ({"bomb", "bmb", "bomb_shooter", "bs", "cannon", "canon", "can", "💣", "boom", "💥"}, "bomb"),
            ({"tack", "tack_shooter", "tac", "tak", "ta", "tacc"}, "tack"),
            ({"glue", "glue_gunner", "gg", "glu"}, "glue"),
            ({"ice", "ice_monkey", "im"}, "ice"),

            ({"sniper", "sniper_monkey", "sn", "sni", "snip", "snooper", "snipermonkey", "gun"}, "sniper"),
            ({"sub", "monkey_sub", "submarine", "u_boat"}, "sub"),
            ({"bucc", "monkey_buccaneer", "boat", "buc", "buccaneer"}, "bucc"),
            ({"heli", "heli-pilot", "helicopter", "helipilot", "hel"}, "heli"),
            ({"ace", "monkey_ace", "pilot", "plane", "airplane"}, "ace"),
            ({"dartling", "gunner", "dlg", "dartl"}, "dartling"),
            ({"mortar", "mortar_monkey", "mor"}, "mortar"),

            ({"druid", "drood", "dru", "druid_monkey"}, "druid"),
            ({"alch", "alchemist", "alk", "alc"}, "alch"),
            ({"mermonkey", "merm", "mermaid", "mer"}, "merm"),
            ({"ninja", "ninja-monkey", "ninja_monkey", "n", "ninj", "nin"}, "ninja"),
            ({"super", "super-monkey", "sup"}, "super"),
            ({"wizard", "wizard_monkey", "apprentice", "wiz"}, "wiz"),

            ({"beast", "beast_handler", "handler", "bh", "breast"}, "beast"),
            ({"engi", "engineer_monkey", "engineer", "engie", "eng", "engie"}, "engi"),
            ({"village", "monkey_village", "vil", "vill", "mvil"}, "vill"),
            ({"spact", "factory", "spike", "spac", "spak", "spanc", "spikes", "spi", "spk", "sf", "spacc", "spike_shooter", "spactory"}, "spac"),
            
            ({'ad', 'ado', 'ador', 'dora', 'priestess', 'odour', 'odoura'}, 'Adora'),
            ({'dj', 'ben', 'benj', 'benny', 'boi', 'benjammin', 'yeet', 'boy'}, 'Benjamin'),
            ({'admiral-brickell', 'admiral', 'brick', 'bri'}, 'Brickell'),
            ({'corv', 'crv', 'crow', 'caw'}, 'Corvus'),
            ({'captain-churchill', 'chirch', 'church', 'hill', 'captain', 'tank', 'winston', 'cap'}, 'Churchill'),
            ({'et', 'eti'}, 'Etienne'),
            ({'voodo', 'vm', 'ezi', 'ezil', 'voodoo'}, 'Ezili'),
            ({'ger', 'gerry', 'morshu', 'gerald'}, 'Geraldo'),
            ({'gwendolin', 'gwendolyn', 'scientist', 'gwend', 'gwendo', 'fire'}, 'Gwen'),
            ({'striker_jones', 'sj', 'striker', 'bones', 'biker', 'stroker'}, 'Jones'),
            ({'obyn-greenfoot', 'obiwan', 'obi', 'oby', 'obeyn'}, 'Obyn'),
            ({'pat-fusty', 'pf', 'fu', 'fusty', 'fro', 'snowman'}, 'Pat'),
            ({'psy', 'Ψ', 'sigh'}, 'Psi'),
            ({'q', 'quin', 'cyberquincy', 'quincey', 'quinc', 'quonc', 'cyber', 'cq'}, 'Quincy'),
            ({'ros', 'rosa', 'rosalina', 'lia'}, 'Rosalia'),
            ({'saud', 'sau'}, 'Sauda')
            
            
            ] for alias in aliases_group[0] for value in [aliases_group[1]]}
aliases.update({}) 

# Plik Excel
EXCEL_FILE = "CRC_database.xlsx"

# Inicjalizacja bota
intents = discord.Intents.default()
intents.message_content = True
bot = discord.Bot(intents=intents)
bot.active_submissions = {}

def calculate_hero_cost(hero_name, level):
    """
    Oblicza całkowity koszt herosa na podstawie poziomu.
    Tylko koszt bazowy herosa jest mnożony przez 1.08 i zaokrąglany do najbliższej 5.
    Koszt poziomów jest mnożony przez levelModifier i zaokrąglany do 1.
    """
    
    # Normalizacja nazwy herosa na podstawie aliasów
    hero_name = hero_name.lower()  # Zamień na małe litery
    hero_name = aliases.get(hero_name, hero_name)  # Zamień alias na podstawową nazwę


    if hero_name not in heroes:  # Upewnij się, że nazwa jest małymi literami
        return 0  # Jeśli heros nie istnieje, zwróć 0

    hero_data = heroes[hero_name]  # Użyj znormalizowanej nazwy
    base_cost = hero_data["cost"]
    level_modifier = hero_data["levelModifier"]

    # Modyfikacja kosztu bazowego: mnożenie przez 1.08 i zaokrąglenie do najbliższej 5
    modified_base_cost = round(base_cost * 1.08 / 5) * 5

    # Koszt poziomów (mnożenie przez levelModifier i zaokrąglenie do 1)
    total_level_cost = 0
    for i in range(1, level):  # Poziomy od 1 do (level - 1)
        level_cost = hero_level_costs[i] * level_modifier
        total_level_cost += round(level_cost)  # Zaokrąglenie do 1

    # Całkowity koszt to zmodyfikowany koszt bazowy + koszt poziomów
    total_cost = modified_base_cost + total_level_cost

    
    return total_cost



def load_excel():
    """
    Wczytuje dane z pliku Excel.
    Jeśli plik nie istnieje, tworzy nowy z odpowiednimi nagłówkami.
    """
    if not os.path.exists(EXCEL_FILE):
        # Tworzenie nowego pliku Excel
        wb = Workbook()
        ws = wb.active
        # Dodaj nagłówki kolumn (rundy)
        ws.append(["Map"] + [f"Round {i}" for i in range(1, 141)])
        # Dodaj wiersze dla każdej mapy
        for map_name in maps.values():
            ws.append([map_name] + [""] * 140)  # Puste komórki dla każdej rundy
        wb.save(EXCEL_FILE)
        return wb, ws

    # Wczytaj istniejący plik
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    return wb, ws

def save_excel(wb):
    """
    Zapisuje dane do pliku Excel.
    """
    wb.save(EXCEL_FILE)

def update_excel(ws, map_name, round_number, price, used_towers, person, version, link):
    """
    Aktualizuje komórkę w pliku Excel dla danej mapy i rundy.
    Dodaje również "Today's Price" na podstawie wartości wież i herosów.
    """
    # Znajdź wiersz dla mapy
    map_row = None
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_col=1, values_only=True), start=2):  # Pomijamy nagłówek
        if row[0] == map_name:
            map_row = row_idx  # Indeks wiersza (liczony od 1)
            break

    if not map_row:
        raise ValueError(f"Map {map_name} not found in Excel file.")

    # Znajdź kolumnę dla rundy
    round_col = round_number + 1  # Pierwsza kolumna to nazwa mapy

    # Przelicz Today's Price na podstawie used_towers
    tower_list = used_towers.split("+")
    total_price = 0
    for tower in tower_list:
        # Sprawdź, czy to heros
        if "-" in tower:
            hero_name, level = tower.split("-")
            level = int(level)
            total_price += calculate_hero_cost(hero_name, level)  # Dodaj koszt herosa
        else:
            # Przetwarzaj wieżę
            parsed_towers, tower_price, error_message = parse_tower(tower)  # Rozpakowanie trzech wartości
            if parsed_towers is None:
                continue  # Pomijaj nieprawidłowe wieże
            total_price += tower_price

    # Zaktualizuj komórkę
    cell_value = f"Price: {price}\nTowers: {used_towers}\nPerson: {person}\nVersion: {version}\nLink: {link}\nToday's Price: {total_price}"
    ws.cell(row=map_row, column=round_col, value=cell_value)


def normalize_map_name(map_name):
    """
    Normalizuje nazwę mapy na podstawie aliasów.
    """
    map_name = map_name.lower().strip()  # Ignoruj wielkość liter i białe znaki
    for map_data in maps_with_aliases.values():
        if map_name == map_data["value"].lower() or map_name in [alias.lower() for alias in map_data["aliases"]]:
            return map_data["value"]  # Zwróć podstawową nazwę mapy
    return None  # Jeśli mapa nie została znaleziona

# Funkcja do normalizacji nazwy wieży
def normalize_tower_name(tower):
    """
    Normalizuje nazwę wieży, ignorując herosy.
    """
    if "#" in tower:
        name, digits = tower.split("#")
    else:
        name, digits = tower, "000"  # Domyślna wartość cyfr

    # Normalizacja nazwy (ignorowanie wielkości liter i zamiana aliasów)
    name = name.lower().strip()  # Ignorowanie wielkości liter i białych znaków
    name = aliases.get(name, name)  # Zamiana aliasów na podstawowe nazwy

    # Połącz nazwę i cyfry z powrotem
    return f"{name}#{digits}"


# /Todo
@bot.slash_command(name="todo", description="Shows TODO list for this bot")
async def todo(ctx: discord.ApplicationContext):
    # Lista rzeczy do zrobienia
    todo_list = """
    **Stuff to do with bot:**
    - somehow copy .xlsx file into spreadsheet
    - add geraldo items
    """

    # Wysłanie wiadomości z listą
    await ctx.respond(todo_list)


def parse_hero(hero):
    """
    Przetwarza herosa w formacie np. Quincy-5.
    Zwraca None (bo heros nie jest wieżą) oraz koszt herosa.
    """
    try:
        hero_name, level = hero.split("-")
        level = int(level)
        
        # Normalizacja nazwy herosa na podstawie aliasów
        hero_name = hero_name.lower()  # Zamień na małe litery
        hero_name = aliases.get(hero_name, hero_name)  # Zamień alias na podstawową nazwę


        # Sprawdź, czy heros istnieje w słowniku heroes
        if hero_name not in heroes:
            return None, 0
        
        # Oblicz koszt herosa
        hero_cost = calculate_hero_cost(hero_name, level)

        return None, hero_cost  # Zwróć koszt herosa
    except (ValueError, KeyError) as e:
        return None, 0

@bot.slash_command(name="crc", description="Display CRC which you want")
async def CRC(
    ctx: discord.ApplicationContext,
    round: discord.Option(int, description="Round (1-140)", required=False),
    map: discord.Option(str, description="Map's name (type manually)", required=False)  # Użytkownik wpisuje nazwę mapy
):
    # Normalizuj nazwę mapy, jeśli została podana
    if map:
        normalized_map = normalize_map_name(map)
        if not normalized_map:
            await ctx.respond(f"Invalid map name. Please choose from: {', '.join(maps_with_aliases.keys())}", ephemeral=True)
            return
        map = normalized_map

    # Reszta logiki komendy /CRC
    wb, ws = load_excel()

    if round is not None and map is not None:
        # Wyświetl wszystkie dane dla konkretnej mapy i rundy
        map_row = None
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_col=1, values_only=True), start=2):  # Pomijamy nagłówek
            if row[0] == map:
                map_row = row_idx  # Indeks wiersza (liczony od 1)
                break

        if not map_row:
            await ctx.respond(f"No data found for Map {map}.")
            return

        round_col = round + 1  # Pierwsza kolumna to nazwa mapy
        cell_value = ws.cell(row=map_row, column=round_col).value

        if not cell_value:
            await ctx.respond(f"No CRC data found for Round {round} on Map {map}.")
            return

        # Konwertuj wartość komórki na ciąg znaków
        cell_value = str(cell_value)

        # Wyodrębnij link z komórki
        link = None
        for line in cell_value.split("\n"):
            if line.startswith("Link:"):
                link = line.split("Link: ")[1]
                break

        # Wyświetl wszystkie dane z komórki (bez linku)
        embed_data = "\n".join([line for line in cell_value.split("\n") if not line.startswith("Link:")])
        embed = discord.Embed(
            title=f"CRC Data for Round {round} on {map}",
            description=embed_data,
            color=discord.Color.blue()
        )
        await ctx.respond(embed=embed)

        # Wyświetl link w osobnej wiadomości
        if link:
            await ctx.send(f"Link: {link}")
        else:
            await ctx.send("No link found in the submission.")

    elif round is not None:
        # Wyświetl dane dla konkretnej rundy (wszystkie mapy)
        round_col = round + 1
        data = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_col=1, values_only=True), start=2):  # Pomijamy nagłówek
            map_name = row[0]
            cell_value = ws.cell(row=row_idx, column=round_col).value
            if cell_value:
                # Konwertuj wartość komórki na ciąg znaków
                cell_value = str(cell_value)

                # Pobierz "Today's Price" z komórki
                today_price = None
                for line in cell_value.split("\n"):
                    if line.startswith("Today's Price:"):
                        today_price = line.split(": ")[1]
                        break
                if today_price:
                    data.append(f"**{map_name}**: {today_price}")

        if not data:
            await ctx.respond(f"No CRC data found for Round {round}.")
            return

        embed = discord.Embed(
            title=f"CRC Data for Round {round}",
            description="\n".join(data),
            color=discord.Color.green()
        )
        await ctx.respond(embed=embed)

    elif map is not None:
        # Wyświetl dane dla konkretnej mapy (wszystkie rundy)
        map_row = None
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_col=1, values_only=True), start=2):  # Pomijamy nagłówek
            if row[0] == map:
                map_row = row_idx  # Indeks wiersza (liczony od 1)
                break

        if not map_row:
            await ctx.respond(f"No data found for Map {map}.")
            return

        data = []
        for round_number in range(1, 141):
            cell_value = ws.cell(row=map_row, column=round_number + 1).value
            if cell_value:
                # Konwertuj wartość komórki na ciąg znaków
                cell_value = str(cell_value)

                # Pobierz "Today's Price" z komórki
                today_price = None
                for line in cell_value.split("\n"):
                    if line.startswith("Today's Price:"):
                        today_price = line.split(": ")[1]
                        break
                if today_price:
                    data.append(f"**Round {round_number}**: {today_price}")

        if not data:
            await ctx.respond(f"No CRC data found for Map {map}.")
            return

        embed = discord.Embed(
            title=f"CRC Data for {map}",
            description="\n".join(data),
            color=discord.Color.orange()
        )
        await ctx.respond(embed=embed)

    else:
        # Brak podanych wartości
        embed = discord.Embed(
            title="Error",
            description="Please provide at least one value (Round or Map).",
            color=discord.Color.red()
        )
        await ctx.respond(embed=embed)

# Modyfikacja funkcji parse_tower
def calculate_single_tower_cost(tower_key):
    """Oblicza koszt pojedynczej wieży z uwzględnieniem modyfikatorów"""
    base_cost = towers.get(tower_key, 0)
    if base_cost == 0:
        return 0
    
    # 1. Pomnóż przez 1.08
    multiplied = base_cost * 1.08
    # 2. Zastosuj floor (usuń część ułamkową)
    floored = math.floor(multiplied)
    # 3. Zaokrąglij do najbliższej wielokrotności 5
    rounded = round(floored / 5) * 5
    
    return rounded

def parse_tower(tower):
    """Rozbija wieżę na wszystkie pośrednie ulepszenia z pełnym przeliczaniem kosztów"""
    try:
        # Normalizacja nazwy
        normalized = normalize_tower_name(tower)
        if not normalized:
            return None, 0, f"Nieznana wieża: {tower}"

        base_name, digits = normalized.split("#")
        if len(digits) != 3:
            return None, 0, "Format musi być nazwa#XYZ (3 cyfry)"

        # Generuj wszystkie wymagane ulepszenia
        components = ["000"]  # Zawsze zaczynamy od podstawowej wersji
        
        for i, digit in enumerate(map(int, digits)):
            if digit < 0 or digit > 5:
                return None, 0, f"Nieprawidłowa cyfra {digit} w pozycji {i}"
                
            position = 2 - i  # 0:setki, 1:dziesiatki, 2:jednosci
            for lvl in range(1, digit + 1):
                value = lvl * (10 ** position)
                components.append(f"{value:03d}")

        # Przelicz każdy komponent
        basic_towers = []
        total_price = 0
        
        for comp in components:
            tower_key = f"{base_name}#{comp}"
            if tower_key not in towers:
                return None, 0, f"Brak w słowniku: {tower_key}"
                
            component_cost = calculate_single_tower_cost(tower_key)
            basic_towers.append((tower_key, component_cost))
            total_price += component_cost

        return basic_towers, total_price, None

    except Exception as e:
        return None, 0, f"Błąd przetwarzania: {str(e)}"


def parse_tower_only(tower):
    """Wersja z obsługą mnożników"""
    try:
        # Rozdziel mnożnik
        multiplier = 1
        if "*" in tower:
            parts = tower.split("*")
            if len(parts) != 2:
                return None, 0, "Zły format mnożnika"
            tower_part, mult = parts
            try:
                multiplier = int(mult)
                if multiplier < 1:
                    return None, 0, "Mnożnik musi być ≥1"
            except ValueError:
                return None, 0, "Mnożnik musi być liczbą"
        else:
            tower_part = tower

        # Parsuj podstawową wieżę
        components, price, error = parse_tower(tower_part)
        if error:
            return None, 0, error

        # Zastosuj mnożnik do wyniku
        multiplied_components = []
        for (tower_key, cost) in components:
            multiplied_components.extend([(tower_key, cost)] * multiplier)
        
        return multiplied_components, price * multiplier, None

    except Exception as e:
        return None, 0, f"Błąd przetwarzania: {str(e)}"



def parse_hero(hero):
    """
    Przetwarza herosa w formacie np. Quincy-5.
    Zwraca None (bo heros nie jest wieżą) oraz koszt herosa.
    """
    try:
        hero_name, level = hero.split("-")
        level = int(level)
        return None, calculate_hero_cost(hero_name, level)  # Zwróć koszt herosa
    except (ValueError, KeyError):
        return None, 0


def get_today_price(ws, map_name, round_number):
    """
    Pobiera wartość "Today's Price" z pliku Excel dla danej mapy i rundy.
    """
    # Znajdź wiersz dla mapy
    map_row = None
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_col=1, values_only=True), start=2):  # Pomijamy nagłówek
        if row[0] == map_name:
            map_row = row_idx  # Indeks wiersza (liczony od 1)
            break

    if not map_row:
        return None  # Jeśli mapa nie istnieje, zwróć None

    # Znajdź kolumnę dla rundy
    round_col = round_number + 1  # Pierwsza kolumna to nazwa mapy

    # Pobierz wartość komórki
    cell_value = ws.cell(row=map_row, column=round_col).value
    if not cell_value:
        return None  # Jeśli komórka jest pusta, zwróć None

    # Przeanalizuj zawartość komórki, aby znaleźć "Today's Price"
    for line in cell_value.split("\n"):
        if line.startswith("Today's Price:"):
            return int(line.split(": ")[1])  # Zwróć wartość "Today's Price"

    return None  # Jeśli nie znaleziono "Today's Price", zwróć None

# Modyfikacja logiki wysyłania wiadomości
async def send_submission_to_channel(channel, round, map, price, used_towers, link, person, version):
    """
    Wysyła zgłoszenie do określonego kanału.
    """
    # Normalizacja nazwy mapy
    normalized_map = normalize_map_name(map)
    if not normalized_map:
        return

    # Rozdziel wieże i herosy
    tower_list = used_towers.split("+")
    normalized_towers = []

    for tower in tower_list:
        if "-" in tower:
            # To jest heros, normalizuj nazwę
            hero_name, level = tower.split("-")
            hero_name = aliases.get(hero_name.lower(), hero_name.lower())  # Zamień alias na podstawową nazwę
            normalized_towers.append(f"{hero_name}-{level}")  # Dodaj znormalizowaną nazwę herosa
        else:
            # To jest wieża, normalizuj
            normalized_towers.append(normalize_tower_name(tower))

    normalized_used_towers = "+".join(normalized_towers)

    # Tworzenie embeda
    embed = discord.Embed(
        title="New CRC Submission",  # Tytuł embeda
        color=discord.Color.blue()   # Kolor embeda
    )
    embed.add_field(name="Round", value=round, inline=True)
    embed.add_field(name="Map", value=normalized_map, inline=True)  # Użyj znormalizowanej nazwy mapy
    embed.add_field(name="Price", value=price, inline=True)
    embed.add_field(name="Used Towers", value=normalized_used_towers, inline=True)
    embed.add_field(name="Person", value=person, inline=True)
    embed.add_field(name="Version", value=version if version else "Unknown", inline=True)

    # Wysłanie embeda do kanału
    submission_channel = bot.get_channel(YOUR_CHANNEL_ID)  # Użyj ID kanału do zgłoszeń
    embed_message = await submission_channel.send(embed=embed)

    # Wysłanie linku jako osobnej wiadomości
    link_message = await submission_channel.send(f"Link: {link}")

    # Dodanie reakcji do wiadomości z embedem
    await embed_message.add_reaction("✅")  # Emotka potwierdzenia
    await embed_message.add_reaction("❌")  # Emotka odrzucenia

    # Przechowanie zgłoszenia w słowniku
    submission_data = {
        "embed_message_id": embed_message.id,  # Tylko ID wiadomości
        "link_message_id": link_message.id,    # Tylko ID wiadomości
        "channel_id": submission_channel.id,   # ID kanału
        "original_user": person,  # Przechowujemy nazwę użytkownika, który wysłał zgłoszenie
        "map": normalized_map,
        "round": round,
        "price": price,
        "used_towers": normalized_used_towers,
        "version": version,
        "link": link
    }
    bot.active_submissions[embed_message.id] = submission_data

# Modyfikacja komendy /submit
@bot.slash_command(name="submit", description="Add new CRC")
async def Submit(
    ctx: discord.ApplicationContext,
    round: discord.Option(int, description="Round (1-140)"),
    map: discord.Option(str, description="Map's name (type manually)"),
    price: discord.Option(int, description="Price"),
    used_towers: discord.Option(str, description="Used Towers (only + works, no *)"),
    link: discord.Option(str, description="Link"),
    person: discord.Option(str, description="Person which did it"),
    version: discord.Option(str, description="Version", required=False)
):
    try:
        # ID specyficznego serwera
        SPECIFIC_SERVER_ID = 1310038138219135047

        # Sprawdź, czy komenda jest używana na odpowiednim serwerze
        if ctx.guild.id != SPECIFIC_SERVER_ID:
            embed = discord.Embed(
                title="Error",
                description="This command is only available on a specific server. To join to server type /invite",
                color=discord.Color.red()
            )
            await ctx.respond(embed=embed, ephemeral=True)
            return

        # Reszta logiki komendy
        if "*" in used_towers:
            embed = discord.Embed(
                title="Error",
                description="The '*' symbol is not allowed in 'used_towers'. Please use only '+' to separate towers.",
                color=discord.Color.red()
            )
            await ctx.respond(embed=embed, ephemeral=True)
            return

        normalized_map = normalize_map_name(map)
        if not normalized_map:
            embed = discord.Embed(
                title="Error",
                description=f"Invalid map name. Please choose from: {', '.join(maps_with_aliases.keys())}",
                color=discord.Color.red()
            )
            await ctx.respond(embed=embed, ephemeral=True)
            return

        global global_version
        submission_version = version if version else global_version

        # Sprawdź wersję tylko jeśli jest podana
        if version and version != global_version:
            await send_submission_to_channel(ctx.channel, round, normalized_map, price, used_towers, link, person, submission_version)
            
            confirm_embed = discord.Embed(
                title="Submission Sent (Different Version)",
                description=f"Your CRC submission has been sent for approval.\nNote: Version {version} differs from global version {global_version} - price validation skipped.",
                color=discord.Color.orange()
            )
            await ctx.respond(embed=confirm_embed)
            return

        # Reszta normalnej logiki sprawdzania cen
        if round < 1 or round > 140:
            embed = discord.Embed(
                title="Error",
                description="Round must be between 1 and 140.",
                color=discord.Color.red()
            )
            await ctx.respond(embed=embed, ephemeral=True)
            return

        tower_list = used_towers.split("+")
        basic_towers = []
        total_price = 0
        unrecognized_towers = []

        for tower in tower_list:
            tower = tower.replace(" ", "").lower()
            if "-" in tower:
                try:
                    hero_name, level = tower.split("-")
                    level = int(level)
                    hero_cost = calculate_hero_cost(hero_name, level)
                    total_price += hero_cost
                    basic_towers.append(f"{hero_name}-{level}")
                except Exception as e:
                    unrecognized_towers.append(f"{tower} (Invalid hero format: {str(e)})")
            else:
                parsed_towers, tower_price, error_message = parse_tower(tower)
                if parsed_towers is None:
                    unrecognized_towers.append(f"{tower} ({error_message})")
                    continue
                basic_towers.extend(parsed_towers)
                total_price += tower_price

        if unrecognized_towers:
            embed = discord.Embed(
                title="Error",
                description=f"The following towers were not recognized:\n{', '.join(unrecognized_towers)}",
                color=discord.Color.red()
            )
            await ctx.respond(embed=embed, ephemeral=True)
            return

        if total_price != price:
            embed = discord.Embed(
                title="Error",
                description=f"Price does not match the value of used towers. Calculated price: {total_price}, provided price: {price}.",
                color=discord.Color.red()
            )
            await ctx.respond(embed=embed, ephemeral=True)
            return

        wb, ws = load_excel()
        today_price = get_today_price(ws, normalized_map, round)

        if today_price is not None and price >= today_price:
            embed = discord.Embed(
                title="Submission Rejected",
                description=f"Your submission for Round {round} on {normalized_map} is too expensive. Today's Price: {today_price}, your price: {price}.",
                color=discord.Color.red()
            )
            await ctx.respond(embed=embed, ephemeral=True)
            return

        await send_submission_to_channel(ctx.channel, round, normalized_map, price, used_towers, link, person, submission_version)

        confirm_embed = discord.Embed(
            title="Submission Sent",
            description="Your CRC submission has been sent for approval.",
            color=discord.Color.green()
        )
        await ctx.respond(embed=confirm_embed)

    except Exception as e:
        error_embed = discord.Embed(
            title="Unexpected Error",
            description=f"An error occurred while processing your submission: {str(e)}",
            color=discord.Color.red()
        )
        await ctx.respond(embed=error_embed, ephemeral=True)
        print(f"Error in submit command: {traceback.format_exc()}")
    
async def delete_submission_messages(submission):
    """
    Usuwa wiadomości związane z zgłoszeniem.
    """
    try:
        # Pobierz kanał
        channel = bot.get_channel(submission["channel_id"])
        if not channel:
            return

        # Usuń wiadomość z embedem
        if "embed_message_id" in submission:
            embed_message = await channel.fetch_message(submission["embed_message_id"])
            await embed_message.delete()

        # Usuń wiadomość z linkiem
        if "link_message_id" in submission:
            link_message = await channel.fetch_message(submission["link_message_id"])
            await link_message.delete()
    except discord.NotFound:
        None
    except discord.Forbidden:
        None
    except Exception as e:
        None


# Obsługa reakcji
@bot.event
async def on_reaction_add(reaction, user):
    """
    Obsługuje reakcje na wiadomości z zgłoszeniami.
    """
    # Sprawdź, czy reakcja została dodana do wiadomości z embedem
    if reaction.message.id in bot.active_submissions:
        submission = bot.active_submissions[reaction.message.id]

        # Sprawdź, czy użytkownik ma rolę Admin lub Oskar
        if any(role.name in ["Oskar", "Admin"] for role in user.roles):
            if reaction.emoji == "✅":
                # Zatwierdzenie zgłoszenia
                # Pobierz dane z zgłoszenia
                round = submission["round"]
                map = submission["map"]
                price = submission["price"]
                used_towers = submission["used_towers"]
                person = submission["original_user"]
                version = submission["version"]
                link = submission["link"]

                # Wczytaj plik Excel
                wb, ws = load_excel()

                # Zaktualizuj plik Excel
                update_excel(ws, map, round, price, used_towers, person, version, link)
                save_excel(wb)

                # Potwierdzenie zatwierdzenia zgłoszenia
                await reaction.message.channel.send(f"Submission from {person} was accepted by {user.name}.")

                # Usunięcie wiadomości z kanału zgłoszeń
                await delete_submission_messages(submission)

                # Usunięcie zgłoszenia ze słownika
                del bot.active_submissions[reaction.message.id]

            elif reaction.emoji == "❌":
                # Odrzucenie zgłoszenia
                person = submission["original_user"]
                await reaction.message.channel.send(f"Submission from {person} was rejected by {user.name}.")

                # Usunięcie wiadomości z kanału zgłoszeń
                await delete_submission_messages(submission)

                # Usunięcie zgłoszenia ze słownika
                del bot.active_submissions[reaction.message.id]

# Komenda /Change_Version
@bot.slash_command(name="change_version", description="Setup version for CRC's (refreshes **Today's Price** based on used towers)")
async def Change_Version(
    ctx: discord.ApplicationContext,
    version: discord.Option(str, description="Towers calculated")
):
    # Sprawdź, czy użytkownik to Oskar4087
    if ctx.author.name != "oskar4807":
        await ctx.respond("Only oskar4807 can use this command.", ephemeral=True)
        return

    global global_version
    global_version = version  # Zaktualizuj globalną wersję

    # Potwierdzenie
    embed = discord.Embed(
        title="Global Version Updated",
        description=f"Global version has been updated to {version}.",
        color=discord.Color.green()
    )
    await ctx.respond(embed=embed)

    # Zapisz zmiany do pliku
    save_excel(wb)

    # Potwierdzenie
    embed = discord.Embed(
        title="Prices Updated",
        description="Today's Price has been recalculated for all entries.",
        color=discord.Color.green()
    )
    await ctx.respond(embed=embed)

@bot.slash_command(name="invite", description="Sends invite to add the bot, and to the Mauler server")
async def links(ctx: discord.ApplicationContext):
    # Link do zaproszenia bota
    bot_invite_link = "https://discord.com/oauth2/authorize?client_id=1347879484711370863&permissions=8798240615489&integration_type=0&scope=applications.commands+bot"
    
    # Link do serwera Discord
    server_invite_link = "https://discord.gg/fr7EWB4"

    # Tworzenie embeda z linkami
    embed = discord.Embed(
        title="Usefull links",
        description="These are links to add the bot and to join to our server:",
        color=discord.Color.blue()
    )
    embed.add_field(name="Add bot", value=f"[Click here]({bot_invite_link})", inline=False)
    embed.add_field(name="Join our server", value=f"[Click here]({server_invite_link})", inline=False)

    # Wysłanie embeda
    await ctx.respond(embed=embed)

def save_submissions(submissions):
    """
    Zapisuje zgłoszenia do pliku JSON.
    """
    with open(SUBMISSIONS_FILE, "w") as f:
        json.dump(submissions, f, indent=4)

def load_submissions():
    """
    Wczytuje zgłoszenia z pliku JSON.
    Jeśli plik nie istnieje, zwraca pusty słownik.
    """
    if not os.path.exists(SUBMISSIONS_FILE):
        return {}
    with open(SUBMISSIONS_FILE, "r") as f:
        return json.load(f)



@bot.slash_command(name="imgur", description="Upload an image to Imgur and get a link")
async def imgur(ctx, image: discord.Option(discord.Attachment, description="Image to upload")):
    # Sprawdź, czy załącznik jest obrazem
    if not image.content_type.startswith('image'):
        await ctx.respond("Please upload a valid image file.", ephemeral=True)
        return

    # Pobierz obraz
    image_data = await image.read()

    # Przygotuj nagłówki dla Imgur API
    headers = {
        "Authorization": f"Client-ID {IMGUR_CLIENT_ID}"
    }

    # Przygotuj dane do wysłania
    files = {
        'image': BytesIO(image_data)
    }

    # Wyślij obraz do Imgur
    response = requests.post(IMGUR_UPLOAD_URL, headers=headers, files=files)

    # Sprawdź odpowiedź
    if response.status_code == 200:
        imgur_data = response.json()
        imgur_link = imgur_data['data']['link']
        await ctx.respond(f"Imgur link: {imgur_link}")
    else:
        await ctx.respond("Failed to upload image to Imgur.", ephemeral=True)



@bot.slash_command(name="aliases", description="Show all available aliases for towers and heroes")
async def aliases_command(ctx: discord.ApplicationContext):
    # Grupowanie aliasów według podstawowych nazw
    alias_groups = {}
    for alias, value in aliases.items():
        if value not in alias_groups:
            alias_groups[value] = []
        alias_groups[value].append(alias)

    # Tworzenie listy aliasów w formacie "value -> alias1, alias2, ..."
    alias_list = []
    for value, aliases_group in alias_groups.items():
        alias_list.append(f"**{value}** → {', '.join(aliases_group)}")

    # Podziel listę na części, aby uniknąć przekroczenia limitu znaków w embedzie
    chunk_size = 50  # Liczba grup aliasów na jedną wiadomość
    alias_chunks = [alias_list[i:i + chunk_size] for i in range(0, len(alias_list), chunk_size)]

    # Wysłanie aliasów w kilku wiadomościach (jeśli jest ich dużo)
    for chunk in alias_chunks:
        embed = discord.Embed(
            title="Aliases",
            description="\n".join(chunk),
            color=discord.Color.blue()
        )
        await ctx.respond(embed=embed)



@bot.event
async def on_ready():
    print(f'Bot {bot.user} is ready!')
    print(f'Commands: {[cmd.name for cmd in bot.application_commands]}')

# Uruchomienie bota
bot.run(TOKEN)
